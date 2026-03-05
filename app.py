import streamlit as st
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from zoneinfo import ZoneInfo
import pandas as pd
from io import BytesIO
import re
import zipfile
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# CONFIG
login_url = "http://sigof.distriluz.com.pe/plus/usuario/login"
FILE_ID = "1td-2WGFN0FUlas0Vx8yYUSb7EZc7MbGWjHDtJYhEY-0"
headers = {
    "User-Agent": "Mozilla/5.0",
    "Referer": login_url,
}

# =========================
# UNIDADES SIGOF
# =========================
UNIDADES = {
    "Ayacucho": 76,
    "Huancayo": 77,
    "Huancavelica": 78,
    "Tarma": 79,
    "Selva Central": 80,
    "Pasco": 81,
    "Huánuco": 82,
    "Valle Mantaro": 83,
    "Tingo María": 84,    
}

CAMBIAR_UNIDAD_URL = "http://sigof.distriluz.com.pe/plus/usuario/ajax_cambiar_sesion"

# PERIODO PERMITIDO
PERIODO_PERMITIDO = "0"

def login_and_get_defecto_iduunn(session, usuario, password):
    credentials = {
        "data[Usuario][usuario]": usuario,
        "data[Usuario][pass]": password
    }
    login_page = session.get(login_url, headers=headers)
    soup = BeautifulSoup(login_page.text, "html.parser")
    csrf_token = soup.find("input", {"name": "_csrf_token"})
    if csrf_token:
        credentials["_csrf_token"] = csrf_token["value"]

    response = session.post(login_url, data=credentials, headers=headers)
    match_iduunn = re.search(r"var DEFECTO_IDUUNN\s*=\s*'(\d+)'", response.text)
    if not match_iduunn:
        return None, False

    defecto_iduunn = int(match_iduunn.group(1))
    dashboard_response = session.get(
        "http://sigof.distriluz.com.pe/plus/dashboard/modulos",
        headers=headers
    )
    if "login" in dashboard_response.text:
        return None, False

    return defecto_iduunn, True

# =========================
# CAMBIO DE UNIDAD SIGOF
# =========================
def cambiar_unidad_sigof(session, iduunn):
    payload = {
        "idempresa": 4,
        "iduunn": iduunn
    }
    session.post(CAMBIAR_UNIDAD_URL, data=payload, headers=headers)
    test = session.get("http://sigof.distriluz.com.pe/plus/dashboard/modulos", headers=headers)
    return str(iduunn) in test.text

def download_excel_from_drive(file_id):
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    response = requests.get(url)
    return pd.read_excel(BytesIO(response.content)) if response.status_code == 200 else None

def descargar_archivo(session, codigo, periodo, nombre_ciclo=None):
    zona = ZoneInfo("America/Lima")
    hoy = datetime.now(zona).strftime("%Y-%m-%d")
    url = (f"http://sigof.distriluz.com.pe/plus/Reportes/ajax_ordenes_historico_xls/U/{hoy}/{hoy}/0/{codigo}/0/0/0/0/0/0/0/0/9/{periodo}")
    response = session.get(url, headers=headers)

    if response.headers.get("Content-Type") == \
       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return response.content, f"{nombre_ciclo}_{periodo}.xlsx" if nombre_ciclo else "archivo.xlsx"
    else:
        return None, None

def main():
    st.set_page_config(page_title="Lmc Lectura", layout="wide")

    st.markdown("""
    <style>
    .stApp {
        background-image: url("https://i.ibb.co/CpVsF4Km/Lecturador.png");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
        height: 100vh;
        position: relative;
    }

    .stApp::before {
        content: "";
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background: linear-gradient(
            rgba(0, 0, 0, 0.75),
            rgba(0, 0, 0, 0.85)
        );
        z-index: 0;
    }

    .stApp > header,
    .stApp > div,
    .main {
        position: relative;
        z-index: 1;
    }

    .main, .block-container {
        width: 100% !important;
        max-width: 100% !important;
        padding: 2rem !important;
    }

    h3 {
        color: #00E0E0 !important;
        text-shadow: 2px 2px 10px rgba(0,0,0,0.8);
    }
    .stTextInput > label,
    .stMultiSelect > label,
    .stSelectbox > label {
        background-color: #595959;
        color: white !important;
        padding: 0.5px 6px;
        border-radius: 8px;
        display: inline-block;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="display: flex; justify-content: center;">
        <h3 style="color:#0078D7;">
            🤖 Control de Avance de Lecturas y Relecturas
        </h3>
    </div>
    """, unsafe_allow_html=True)

    if "session" not in st.session_state:
        st.session_state.session = None
    if "defecto_iduunn" not in st.session_state:
        st.session_state.defecto_iduunn = None
    if "ciclos_disponibles" not in st.session_state:
        st.session_state.ciclos_disponibles = {}
    if "archivos_descargados" not in st.session_state:
        st.session_state.archivos_descargados = {}
    if "mostrar_resumen" not in st.session_state:
        st.session_state.mostrar_resumen = True

    # ===== LOGIN =====
    if st.session_state.session is None:
        usuario = st.text_input("🤵 Humano ingrese su usuario sigof",
                                placeholder="Usuario sigof", max_chars=20)
        password = st.text_input("🔑 Humano ingrese su contraseña sigof",
                                 placeholder="Contraseña sigof",
                                 type="password", max_chars=26)

        if st.button("🔓 Humano inicie sesión"):
            session = requests.Session()
            defecto_iduunn, ok = login_and_get_defecto_iduunn(session, usuario, password)

            if not ok:
                st.error("❌ Humano error de autenticación")
                return

            st.session_state.session = session
            st.session_state.defecto_iduunn = defecto_iduunn
            st.session_state.mostrar_resumen = True

            df = download_excel_from_drive(FILE_ID)
            df["id_unidad"] = pd.to_numeric(df["id_unidad"], errors="coerce")
            df = df[df["id_unidad"] == defecto_iduunn]

            st.session_state.ciclos_disponibles = {
                f"{r['Id_ciclo']} {r['nombre_ciclo']}": str(r["Id_ciclo"])
                for _, r in df.iterrows()
            }
            st.rerun()

    # ======================
    # SELECCIÓN DE UNIDAD
    # ======================
    if st.session_state.session is not None:
        nombre_actual = {v: k for k, v in UNIDADES.items()}.get(st.session_state.defecto_iduunn, "Huancayo")

        unidad = st.selectbox("🏢 Humano seleccione su unidad empresarial o operativa",
                              list(UNIDADES.keys()),
                              index=list(UNIDADES.keys()).index(nombre_actual))

        if st.button("🔄 Cambiar Unidad"):
            nuevo = UNIDADES[unidad]

            if nuevo != st.session_state.defecto_iduunn:
                ok = cambiar_unidad_sigof(st.session_state.session, nuevo)
                if not ok:
                    st.error("❌ Humano SIGOF rechazó el cambio de unidad")
                    st.stop()

                df = download_excel_from_drive(FILE_ID)
                df["id_unidad"] = pd.to_numeric(df["id_unidad"], errors="coerce").fillna(-1).astype(int)
                df = df[df["id_unidad"] == nuevo]

                st.session_state.ciclos_disponibles = {
                    f"{r['Id_ciclo']} {r['nombre_ciclo']}": str(r["Id_ciclo"])
                    for _, r in df.iterrows()
                }

                st.session_state.defecto_iduunn = nuevo
                st.session_state.unidad_actual = nuevo
                st.success(f"Unidad cambiada a {unidad}")
                st.rerun()

    # ===== DESCARGA DE CICLOS =====
    if st.session_state.ciclos_disponibles:
        opciones = list(st.session_state.ciclos_disponibles.keys())

        col1, col2 = st.columns([4, 0.5])
        with col1:
            seleccionados = st.multiselect("🔎 Humano seleccione sus ciclos", opciones)
        with col2:
            periodo = st.text_input("📅 Periodo actual",
                                    value=PERIODO_PERMITIDO,
                                    disabled=True)

        if st.button("📥 Humano Mostrar el % Avance y RL"):
            st.session_state.mostrar_resumen = True

            if not seleccionados:
                st.error("❌ Humano seleccione al menos un ciclo.")
            else:
                st.session_state.archivos_descargados.clear()

                with ThreadPoolExecutor(max_workers=8) as executor:

                    tareas = []
                
                    for nombre in seleccionados:
                        codigo = st.session_state.ciclos_disponibles[nombre]
                
                        tareas.append(
                            executor.submit(
                                descargar_archivo,
                                st.session_state.session,
                                codigo,
                                periodo,
                                nombre
                            )
                        )
                
                    for future in as_completed(tareas):
                
                        contenido, filename = future.result()
                
                        if contenido:
                            st.session_state.archivos_descargados[filename] = contenido

                if not st.session_state.archivos_descargados:
                    st.markdown(
                        "<p style='color:red; font-weight:bold; font-size:16px;'>"
                        "Humano, la información ya no está disponible en SIGOF WEB "
                        "debido a que el periodo de lectura ha finalizado. "
                        "Por favor, espere el siguiente periodo."
                        "</p>",
                        unsafe_allow_html=True
                    )
                    st.session_state.mostrar_resumen = False

        # ===== RESUMEN =====
        if st.session_state.archivos_descargados and st.session_state.mostrar_resumen:
            st.markdown(
                "<p style='font-size:20px; color:#00E0E0;'>👷 RESUMEN POR LECTURADOR</p>",
                unsafe_allow_html=True
            )

            resumen_total = []
            st.session_state.detalle_relecturas_global = pd.DataFrame()

            for filename, contenido in st.session_state.archivos_descargados.items():
                df_excel = pd.read_excel(BytesIO(contenido))

                if {"lecturista", "resultado"}.issubset(df_excel.columns):

                    resumen = (
                        df_excel
                        .groupby("lecturista")
                        .agg(
                            Asignados=("lecturista", "size"),
                            Avance=("resultado", lambda x: x.notna().sum())
                        )
                        .reset_index()
                    )

                    # ==== % FOTOS ====
                    if "foto" in df_excel.columns:
                        resumen["Fotos"] = (
                            df_excel
                            .groupby("lecturista")["foto"]
                            .apply(lambda x: x.astype(str).str.lower().str.contains("ver foto").sum())
                            .values
                        )
                    else:
                        resumen["Fotos"] = 0

                    resumen["Descargados_temp"] = (
                        df_excel
                        .groupby("lecturista")["resultado"]
                        .apply(lambda x: x.notna().sum())
                        .values
                    )

                    # ==== RELECTURAS (CORREGIDO) ====
                    if {"tipo_lectura", "resultado"}.issubset(df_excel.columns):
                    
                        df_temp = df_excel.copy()

                        # Guardar detalle de relecturas
                        df_temp["es_relectura"] = (
                            (df_temp["tipo_lectura"].astype(str).str.upper().str.strip() == "R") &
                            (
                                df_temp["resultado"].isna() |
                                (df_temp["resultado"].astype(str).str.strip() == "")
                            )
                        )
                        
                        detalle_relecturas = df_temp[df_temp["es_relectura"] == True].copy()
                        
                        # Guardar global
                        if "detalle_relecturas_global" not in st.session_state:
                            st.session_state.detalle_relecturas_global = pd.DataFrame()
                        
                        st.session_state.detalle_relecturas_global = pd.concat(
                            [st.session_state.detalle_relecturas_global, detalle_relecturas],
                            ignore_index=True
                        )

                    
                        df_temp["es_relectura"] = (
                            (df_temp["tipo_lectura"].astype(str).str.upper().str.strip() == "R") &
                            (
                                df_temp["resultado"].isna() |
                                (df_temp["resultado"].astype(str).str.strip() == "")
                            )
                        )
                    
                        relecturas = (
                            df_temp
                            .groupby("lecturista")["es_relectura"]
                            .sum()
                            .reset_index()
                        )
                    
                        resumen = resumen.merge(relecturas, on="lecturista", how="left")
                        resumen["RL"] = resumen["es_relectura"].fillna(0).astype(int)
                        resumen.drop(columns=["es_relectura"], inplace=True)
                    
                    else:
                        resumen["RL"] = 0
        #==========================================================================

                    resumen["Descargados / Finalizados"] = resumen["Descargados_temp"].apply(
                        lambda x: f"{x}" if x > 0 else "0 descargas o no inicia"
                    )

                    resumen.drop(columns=["Descargados_temp"], inplace=True)
                    resumen["Pendientes"] = resumen["Asignados"] - resumen["Avance"]

                    if resumen.empty or resumen["Asignados"].sum() == 0:
                        st.markdown(
                            "<p style='color:red; font-weight:bold; font-size:16px;'>"
                            "Humano, la información ya no está disponible en SIGOF WEB "
                            "debido a que el periodo de reparto ha finalizado. "
                            "Por favor, espere el siguiente periodo."
                            "</p>",
                            unsafe_allow_html=True
                        )
                        st.session_state.mostrar_resumen = False
                        resumen_total.clear()
                        break

                    resumen["% de Avance"] = (resumen["Avance"] / resumen["Asignados"] * 100).round(2)
                    resumen["% de Fotos"] = (resumen["Fotos"] / resumen["Asignados"] * 100).round(2)

                    resumen.drop(columns=["Avance", "Fotos"], inplace=True)
                    resumen["Ciclo"] = filename

                    cols = list(resumen.columns)
                    i = cols.index("% de Avance")
                    cols.insert(i + 1, cols.pop(cols.index("% de Fotos")))
                    resumen = resumen[cols]

                    resumen_total.append(resumen)

            if resumen_total and st.session_state.mostrar_resumen:
                df_final = pd.concat(resumen_total, ignore_index=True)

                df_final = df_final.rename(columns={
                    "lecturista": "Lecturador",
                    "Asignados": "Asig",
                    "Descargados / Finalizados": "Des/Fin",
                    "Pendientes": "Pend",
                    "% de Avance": "% de Avance",
                    "% de Fotos": "% de Fotos",
                    "Ciclo": "Ciclo Lectura"
                })
                # Forzar alineación a la derecha convirtiendo Des/Fin en número cuando sea posible
                #df_final["Des/Fin"] = pd.to_numeric(df_final["Des/Fin"], errors="coerce").fillna(df_final["Des/Fin"])
                df_final["Des/Fin"] = df_final["Des/Fin"].apply(
                    lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace(".", "").isdigit() else x
                )

                # Ocultar columna RL en el resumen (pero mantenerla para el detalle)
                if "RL" in df_final.columns:
                    df_final_sin_rl = df_final.drop(columns=["RL"])
                else:
                    df_final_sin_rl = df_final.copy()

                def color_avance(val):
                    if val < 59:
                        return "color: red; font-weight: bold;"
                    elif 59 <= val <= 90:
                        return "color: orange; font-weight: bold;"
                    else:
                        return "color: green; font-weight: bold;"

                df_styled = (
                    df_final_sin_rl.style
                    .format({"% de Avance": "{:.2f}%", "% de Fotos": "{:.2f}%"})
                    .applymap(color_avance, subset=["% de Avance"])
                )

                st.dataframe(df_styled, use_container_width=True)

                # ===============================
                # DETALLE DE RELECTURAS
                # ===============================
                #st.markdown("### 🔎 Ver detalle de Relecturas")
                if st.session_state.archivos_descargados and st.session_state.mostrar_resumen:
                    st.markdown(
                        "<p style='font-size:20px; color:#00E0E0;'>👷 Lecturistas con Relecturas</p>",
                        unsafe_allow_html=True
                    )
                
                if "detalle_relecturas_global" in st.session_state:
                
                    detalle = st.session_state.detalle_relecturas_global.copy()
                
                    if not detalle.empty:
                
                        # Crear columna RL
                        detalle["RL"] = "R"
                
                        columnas_mostrar = [
                            col for col in 
                            ["ciclo", "sector", "ruta", "RL", "suministro", "lecturista", "cliente", "direccion"]
                            if col in detalle.columns
                        ]
                
                        st.info(f"Total suministros con relecturas: {len(detalle)}")
                
                        st.dataframe(
                            detalle[columnas_mostrar],
                            use_container_width=True
                        )
                    else:
                        st.info("Humano no tiene suministros con relecturas en los ciclos seleccionados.")
                

                # ===== EXPORTACIÓN =====
                buffer_excel = BytesIO()
                with pd.ExcelWriter(buffer_excel, engine="openpyxl") as writer:
                    df_final.to_excel(writer, index=False, sheet_name="Resumen por Lecturista")    
                    workbook = writer.book
                    worksheet = writer.sheets["Resumen por Lecturista"]
    
                    from openpyxl.styles import PatternFill, Font, Border, Alignment
    
                    header_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    no_border = Border()
    
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.border = no_border
    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
    
                    for col in worksheet.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        worksheet.column_dimensions[col_letter].width = max_length + 3
    
                    col_idx = None
                    for i, cell in enumerate(worksheet[1], start=1):
                        if cell.value == "% de Avance":
                            col_idx = i
                            break
    
                    if col_idx:
                        for row in range(2, worksheet.max_row + 1):
                            c = worksheet.cell(row=row, column=col_idx)
                            if isinstance(c.value, (int, float)):
                                c.value = c.value / 100
                                c.number_format = "0.00%"
    
                    last_row = worksheet.max_row + 1
                    last_col = worksheet.max_column
                    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=last_col)
    
                    cell = worksheet.cell(row=last_row, column=1)
                    cell.value = "Data LMC"
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
    
                buffer_excel.seek(0)
    
                st.download_button(
                    "📊 Humano Exportar Excel",
                    data=buffer_excel,
                    file_name="Lmc_Resumen_por_Lecturador👷.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    # ===== CERRAR SESIÓN =====
    if st.session_state.session is not None:
        if st.button("🔒 Cerrar sesión"):
            st.session_state.session = None
            st.session_state.defecto_iduunn = None
            st.session_state.ciclos_disponibles = {}
            st.session_state.archivos_descargados = {}
            st.session_state.mostrar_resumen = True
            st.rerun()

if __name__ == "__main__":
    main()

# --- Footer fijo ---
st.markdown("""
<style>
.footer {
    position: fixed;
    bottom: 0;
    width: 100%;
    background-color: white;
    padding: 10px 8px;
    text-align: center;
    font-size: 15px;
    color: #262626;
    z-index: 9999;
    border-top: 1px solid #ddd;
}
</style>
<div class="footer">Desarrollado por Luis Miguel Cahuana Figueroa.</div>
""", unsafe_allow_html=True)
